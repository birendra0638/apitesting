import requests
import json
import openpyxl


def test_add_mul_students():
    api_url = "http://thetestingworldapi.com/api/studentsDetails"
    f = open("F:/Python_files/API_Test/add_student.json")
    wb = openpyxl.load_workbook("F:/Python_files/API_Test/test_data.xlsx")
    sh = wb['data']
    rows = sh.max_row
    json_request = json.loads(f.read())
    for i in range(2, rows + 1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value

        response = requests.post(api_url, json_request)
        print(response.text)

        #print(response.status_code)
        assert response.status_code == 201
