import requests
import json
import openpyxl
from openpyxl import Workbook

from DataDriv import Lib


def test_add_mul_students():
    api_url = "http://thetestingworldapi.com/api/studentsDetails"
    f = open("F:/Python_files/API_Test/add_student.json")
    json_request = json.loads(f.read())

    obj = Lib.CommonLib('F:/Python_files/API_Test/test_data.xlsx', 'data')
    col = obj.column_count()
    row = obj.row_count()
    print("rowcount", row)
    keyList = obj.key_names()
    wb = Workbook()
    wb['Sheet'].title = "API OP"
    sh = wb.active
    for i in range(2, row + 1):
        updated_json_req = obj.update_request_with_data(i, json_request, keyList)

        response = requests.post(api_url, updated_json_req)
        print(response)
        print(json_request['first_name'].replace('&', 'amp').replace('"', 'quot'))

        sh['A1'].value = "Request"
        sh['B1'].value = "Response"
        # sh['A2'].value = str(json_request)
        # sh['B2'].value = str(response)
        print('value of i', i)
        sh.cell(i, 1).value = str(json_request)
        sh.cell(i, 2).value = str(response)

    wb.save("finalReport.xlsx")
