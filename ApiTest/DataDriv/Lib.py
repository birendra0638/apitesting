import json
import jsonpath
import requests
import openpyxl


class CommonLib:
    def __init__(self, FilePath, SheetName):
        global wb
        global sh
        wb = openpyxl.load_workbook(FilePath)
        sh = wb[SheetName]

    def row_count(self):
        rows = sh.max_row
        return rows

    def column_count(self):
        cols = sh.max_column
        return cols

    def key_names(self):
        c = sh.max_column
        li = []
        for i in range(1, c + 1):
            cell = sh.cell(row=1, column=i)
            li.insert(i - 1, cell.value)
        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sh.max_column
        for i in range(1, c + 1):
            cell = sh.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i - 1]] = cell.value
        return jsonRequest

    def writeData(self,rowNum, colNum, data):
        workbook = openpyxl.load_workbook(wb)
        sheet = workbook[sh]
        sheet.cell(row=rowNum, column=colNum).value = data
        workbook.save(wb)
